from datetime import date
from typing import List, Optional
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract
from fastapi import HTTPException

from app.models.financial_transaction import FinancialTransaction
from app.models.package_request import PackageRequest
from app.models.package import Package
from app.schemas.financial_transaction import (
    FinancialTransactionCreate,
    FinancialTransactionUpdate,
    CategorySummary,
    MonthlyFlow,
    FinancialSummaryResponse,
)

MONTH_NAMES_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]


def list_transactions(
    db: Session,
    type: Optional[str] = None,
    category: Optional[str] = None,
    company_id: Optional[int] = None,
    project_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 200,
) -> List[FinancialTransaction]:
    q = db.query(FinancialTransaction).options(
        joinedload(FinancialTransaction.company),
        joinedload(FinancialTransaction.project),
        joinedload(FinancialTransaction.created_by),
    )

    if type:
        q = q.filter(FinancialTransaction.type == type)
    if category:
        q = q.filter(FinancialTransaction.category == category)
    if company_id:
        q = q.filter(FinancialTransaction.company_id == company_id)
    if project_id:
        q = q.filter(FinancialTransaction.project_id == project_id)
    if start_date:
        q = q.filter(FinancialTransaction.transaction_date >= start_date)
    if end_date:
        q = q.filter(FinancialTransaction.transaction_date <= end_date)
    if search:
        s = f"%{search.lower()}%"
        q = q.filter(
            func.lower(FinancialTransaction.title).like(s) |
            func.lower(FinancialTransaction.category).like(s) |
            func.lower(FinancialTransaction.payment_reference).like(s)
        )

    return q.order_by(FinancialTransaction.transaction_date.desc(), FinancialTransaction.id.desc()).offset(skip).limit(limit).all()


def create_transaction(db: Session, data: FinancialTransactionCreate, creator_id: int) -> FinancialTransaction:
    transaction = FinancialTransaction(
        **data.model_dump(),
        created_by_id=creator_id
    )
    db.add(transaction)
    db.commit()
    db.refresh(transaction)
    return db.query(FinancialTransaction).options(
        joinedload(FinancialTransaction.company),
        joinedload(FinancialTransaction.project),
        joinedload(FinancialTransaction.created_by),
    ).filter(FinancialTransaction.id == transaction.id).first()


def update_transaction(db: Session, transaction_id: int, data: FinancialTransactionUpdate) -> FinancialTransaction:
    transaction = db.query(FinancialTransaction).filter(FinancialTransaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Registro financiero no encontrado")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(transaction, field, value)

    db.commit()
    db.refresh(transaction)
    return db.query(FinancialTransaction).options(
        joinedload(FinancialTransaction.company),
        joinedload(FinancialTransaction.project),
        joinedload(FinancialTransaction.created_by),
    ).filter(FinancialTransaction.id == transaction.id).first()


def delete_transaction(db: Session, transaction_id: int):
    transaction = db.query(FinancialTransaction).filter(FinancialTransaction.id == transaction_id).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Registro financiero no encontrado")
    db.delete(transaction)
    db.commit()


def get_financial_summary(db: Session) -> FinancialSummaryResponse:
    # 1. Total Ingresos Manuales
    ingresos_manuales_res = db.query(
        func.coalesce(func.sum(FinancialTransaction.amount), 0.0),
        func.count(FinancialTransaction.id)
    ).filter(FinancialTransaction.type == "ingreso").first()

    total_ingresos_manuales = float(ingresos_manuales_res[0] if ingresos_manuales_res else 0.0)
    ingresos_count = int(ingresos_manuales_res[1] if ingresos_manuales_res else 0)

    # 2. Total Ingresos QR Verificados (Suscripciones)
    qr_verified_res = db.query(
        func.coalesce(func.sum(Package.base_price), 0.0)
    ).join(PackageRequest, PackageRequest.package_id == Package.id).filter(
        PackageRequest.payment_status == "pago_verificado"
    ).scalar()

    total_ingresos_qr = float(qr_verified_res or 0.0)

    # 3. Total Egresos Globales
    egresos_res = db.query(
        func.coalesce(func.sum(FinancialTransaction.amount), 0.0),
        func.count(FinancialTransaction.id)
    ).filter(FinancialTransaction.type == "egreso").first()

    total_egresos_global = float(egresos_res[0] if egresos_res else 0.0)
    egresos_count = int(egresos_res[1] if egresos_res else 0)

    total_ingresos_global = total_ingresos_manuales + total_ingresos_qr
    balance_neto = total_ingresos_global - total_egresos_global

    # 4. Desglose por categorías
    cat_rows = db.query(
        FinancialTransaction.category,
        FinancialTransaction.type,
        func.coalesce(func.sum(FinancialTransaction.amount), 0.0),
        func.count(FinancialTransaction.id)
    ).group_by(FinancialTransaction.category, FinancialTransaction.type).all()

    categories_breakdown = [
        CategorySummary(
            category=row[0],
            type=row[1],
            total_amount=float(row[2]),
            count=int(row[3])
        )
        for row in cat_rows
    ]

    # 5. Flujo mensual (Últimos 12 meses)
    monthly_rows = db.query(
        extract('year', FinancialTransaction.transaction_date).label('yr'),
        extract('month', FinancialTransaction.transaction_date).label('mo'),
        FinancialTransaction.type,
        func.coalesce(func.sum(FinancialTransaction.amount), 0.0)
    ).group_by('yr', 'mo', FinancialTransaction.type).order_by('yr', 'mo').all()

    monthly_map = {}
    for yr, mo, t_type, amt in monthly_rows:
        yr_i, mo_i = int(yr), int(mo)
        key = (yr_i, mo_i)
        if key not in monthly_map:
            monthly_map[key] = {"ingreso": 0.0, "egreso": 0.0}
        monthly_map[key][t_type] += float(amt)

    monthly_flow = []
    for (y, m), val in sorted(monthly_map.items()):
        month_name = MONTH_NAMES_ES[m] if 1 <= m <= 12 else str(m)
        ing = val["ingreso"]
        eg = val["egreso"]
        monthly_flow.append(MonthlyFlow(
            year=y,
            month=m,
            month_name=f"{month_name} {y}",
            total_ingresos=ing,
            total_egresos=eg,
            balance=ing - eg
        ))

    return FinancialSummaryResponse(
        total_ingresos_manuales=total_ingresos_manuales,
        total_ingresos_qr_verificados=total_ingresos_qr,
        total_ingresos_global=total_ingresos_global,
        total_egresos_global=total_egresos_global,
        balance_neto=balance_neto,
        ingresos_count=ingresos_count,
        egresos_count=egresos_count,
        categories_breakdown=categories_breakdown,
        monthly_flow=monthly_flow
    )


def export_finances_excel(
    db: Session,
    type: Optional[str] = None,
    company_id: Optional[int] = None,
    project_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> StreamingResponse:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    transactions = list_transactions(
        db, type=type, company_id=company_id, project_id=project_id,
        start_date=start_date, end_date=end_date, limit=1000
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    # Título Principal
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "BENCHAMEN - REPORTE DE MOVIMIENTOS FINANCIEROS"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras de Tabla
    headers = [
        "ID", "Tipo", "Concepto / Título", "Monto (Bs.)", "Categoría",
        "Método de Pago", "Nro. Referencia", "Fecha", "Empresa", "Proyecto", "Registrado Por"
    ]
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 20

    ws.column_dimensions["C"].width = 35  # Título más ancho

    total_ingresos = 0.0
    total_egresos = 0.0

    for r_idx, t in enumerate(transactions, 4):
        amt = float(t.amount or 0.0)
        t_type = "INGRESO" if t.type == "ingreso" else "EGRESO"
        if t.type == "ingreso":
            total_ingresos += amt
        else:
            total_egresos += amt

        row_data = [
            t.id,
            t_type,
            t.title,
            amt,
            t.category,
            t.payment_method or "-",
            t.payment_reference or "-",
            str(t.transaction_date),
            t.company.name if t.company else "General",
            t.project.name if t.project else "General",
            t.created_by.name if t.created_by else "-"
        ]

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 2:  # Tipo
                if t.type == "ingreso":
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    cell.font = Font(color="15803D", bold=True)
                else:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    cell.font = Font(color="B91C1C", bold=True)
            elif c_idx == 4:  # Monto
                cell.number_format = "#,##0.00"
                if t.type == "ingreso":
                    cell.font = Font(color="15803D", bold=True)
                else:
                    cell.font = Font(color="B91C1C", bold=True)

    # Fila de Totales
    summary_row = len(transactions) + 5
    ws.cell(row=summary_row, column=3, value="TOTAL INGRESOS:").font = Font(bold=True)
    c_ing = ws.cell(row=summary_row, column=4, value=total_ingresos)
    c_ing.font = Font(bold=True, color="15803D")
    c_ing.number_format = "#,##0.00"

    ws.cell(row=summary_row + 1, column=3, value="TOTAL EGRESOS:").font = Font(bold=True)
    c_eg = ws.cell(row=summary_row + 1, column=4, value=total_egresos)
    c_eg.font = Font(bold=True, color="B91C1C")
    c_eg.number_format = "#,##0.00"

    ws.cell(row=summary_row + 2, column=3, value="BALANCE NETO:").font = Font(bold=True, size=11)
    c_bal = ws.cell(row=summary_row + 2, column=4, value=total_ingresos - total_egresos)
    c_bal.font = Font(bold=True, color="0EA5E9" if total_ingresos >= total_egresos else "B91C1C", size=11)
    c_bal.number_format = "#,##0.00"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_financiero_{date.today()}.xlsx"},
    )
