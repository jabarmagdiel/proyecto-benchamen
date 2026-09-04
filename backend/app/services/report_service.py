import io
from datetime import date
from typing import Optional

from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi.responses import StreamingResponse

from app.models.activity import Activity
from app.models.project import Project
from app.models.company import Company
from app.models.user import User
from app.utils.enums import ActivityStatus

# ─── Diccionarios de traducción a Español ──────────────────────────────────────
ACTIVITY_TYPE_LABELS = {
    "filmacion": "Filmación",
    "edicion_video": "Edición de Video",
    "diseno_grafico": "Diseño Gráfico",
    "fotografia": "Fotografía",
    "copywriting": "Copywriting",
    "publicacion_redes": "Publicación en Redes",
    "planificacion_contenido": "Planificación de Contenido",
    "reunion_cliente": "Reunión con Cliente",
    "entrega_material": "Entrega de Material",
    "otro": "Otro",
}

ACTIVITY_STATUS_LABELS = {
    "pendiente": "Pendiente",
    "asignada": "Asignada",
    "en_proceso": "En Proceso",
    "en_revision": "En Revisión",
    "observada": "Observada",
    "aprobada": "Aprobada",
    "cancelada": "Cancelada",
}

PRIORITY_LABELS = {
    "baja": "Baja",
    "media": "Media",
    "alta": "Alta",
    "urgente": "Urgente",
}


def _get_activities_data(
    db: Session,
    current_user: Optional[User] = None,
    company_id: Optional[int] = None,
    project_id: Optional[int] = None,
    assigned_user_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    q = (
        db.query(
            Activity.id,
            Activity.title,
            Activity.activity_type,
            Activity.priority,
            Activity.status,
            Activity.deadline,
            Activity.created_at,
            Activity.time_spent_seconds,
            User.name.label("assigned_user"),
            Project.name.label("project"),
            Company.name.label("company"),
        )
        .outerjoin(User, User.id == Activity.assigned_user_id)
        .join(Project, Project.id == Activity.project_id)
        .join(Company, Company.id == Project.company_id)
    )
    if current_user and (current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)) == "gerencia":
        user_dept_ids = [d.id for d in current_user.departments]
        if user_dept_ids:
            from app.models.user import user_departments
            dept_user_ids = db.query(user_departments.c.user_id).filter(user_departments.c.department_id.in_(user_dept_ids)).subquery()
            dept_proj_ids = db.query(Project.id).filter(Project.department_id.in_(user_dept_ids)).subquery()
            q = q.filter(
                (Activity.assigned_user_id == current_user.id) |
                (Activity.created_by_id == current_user.id) |
                (Activity.assigned_user_id.in_(dept_user_ids)) |
                (Activity.project_id.in_(dept_proj_ids))
            )
    if company_id:
        q = q.filter(Company.id == company_id)
    if project_id:
        q = q.filter(Activity.project_id == project_id)
    if assigned_user_id:
        q = q.filter(Activity.assigned_user_id == assigned_user_id)
    if status:
        q = q.filter(Activity.status == status)
    if date_from:
        q = q.filter(Activity.created_at >= date_from)
    if date_to:
        q = q.filter(Activity.created_at <= date_to)
    return q.order_by(Activity.created_at.desc()).all()


def get_analytics_report(
    db: Session,
    current_user: User,
    company_id: Optional[int] = None,
    project_id: Optional[int] = None,
    assigned_user_id: Optional[int] = None,
    status: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    from app.models.financial_transaction import FinancialTransaction
    today = date.today()

    # Query base de actividades
    act_q = db.query(Activity).outerjoin(Project, Project.id == Activity.project_id).outerjoin(Company, Company.id == Project.company_id)
    role_str = current_user.role.value if hasattr(current_user.role, 'value') else str(current_user.role)

    if role_str == "gerencia":
        user_dept_ids = [d.id for d in current_user.departments]
        if user_dept_ids:
            from app.models.user import user_departments
            dept_user_ids = db.query(user_departments.c.user_id).filter(user_departments.c.department_id.in_(user_dept_ids)).subquery()
            dept_proj_ids = db.query(Project.id).filter(Project.department_id.in_(user_dept_ids)).subquery()
            act_q = act_q.filter(
                (Activity.assigned_user_id == current_user.id) |
                (Activity.created_by_id == current_user.id) |
                (Activity.assigned_user_id.in_(dept_user_ids)) |
                (Activity.project_id.in_(dept_proj_ids))
            )
    elif role_str == "operativo":
        act_q = act_q.filter(Activity.assigned_user_id == current_user.id)
    elif role_str == "cliente":
        act_q = act_q.filter(Company.id == current_user.company_id)

    if company_id:
        act_q = act_q.filter(Company.id == company_id)
    if project_id:
        act_q = act_q.filter(Activity.project_id == project_id)
    if assigned_user_id:
        act_q = act_q.filter(Activity.assigned_user_id == assigned_user_id)
    if status:
        act_q = act_q.filter(Activity.status == status)
    if date_from:
        act_q = act_q.filter(Activity.created_at >= date_from)
    if date_to:
        act_q = act_q.filter(Activity.created_at <= date_to)

    activities = act_q.all()

    total_activities = len(activities)
    completed_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) == "aprobada")
    pending_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) in ["pendiente", "asignada"])
    in_progress_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) == "en_proceso")
    in_review_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) == "en_revision")
    observed_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) == "observada")
    cancelled_activities = sum(1 for a in activities if (a.status.value if hasattr(a.status, 'value') else str(a.status)) == "cancelada")
    late_activities = sum(1 for a in activities if a.deadline and a.deadline < today and (a.status.value if hasattr(a.status, 'value') else str(a.status)) not in ["aprobada", "cancelada"])

    total_time_seconds = sum(a.time_spent_seconds or 0 for a in activities)
    completion_rate = round((completed_activities / total_activities * 100), 1) if total_activities > 0 else 0.0

    # Desglose por estado
    status_dict = {}
    for a in activities:
        st_key = a.status.value if hasattr(a.status, 'value') else str(a.status)
        status_dict[st_key] = status_dict.get(st_key, 0) + 1
    
    activity_by_status = [
        {"status": k, "name": ACTIVITY_STATUS_LABELS.get(k, k), "count": v}
        for k, v in status_dict.items()
    ]

    # Desglose por tipo de actividad
    type_dict = {}
    for a in activities:
        tp_key = a.activity_type.value if hasattr(a.activity_type, 'value') else str(a.activity_type)
        type_dict[tp_key] = type_dict.get(tp_key, 0) + 1

    activity_by_type = [
        {"type": k, "name": ACTIVITY_TYPE_LABELS.get(k, k), "count": v}
        for k, v in type_dict.items()
    ]

    # Rendimiento por Persona/Usuario
    user_stats = {}
    for a in activities:
        uid = a.assigned_user_id or 0
        uname = a.assigned_user.name if a.assigned_user else "Sin asignar"
        if uid not in user_stats:
            user_stats[uid] = {"user_id": uid, "user_name": uname, "total": 0, "completed": 0, "late": 0, "time_seconds": 0}
        
        user_stats[uid]["total"] += 1
        st_val = a.status.value if hasattr(a.status, 'value') else str(a.status)
        if st_val == "aprobada":
            user_stats[uid]["completed"] += 1
        if a.deadline and a.deadline < today and st_val not in ["aprobada", "cancelada"]:
            user_stats[uid]["late"] += 1
        user_stats[uid]["time_seconds"] += (a.time_spent_seconds or 0)

    user_performance = []
    for u in user_stats.values():
        eff = round((u["completed"] / u["total"] * 100), 1) if u["total"] > 0 else 0.0
        user_performance.append({
            "user_id": u["user_id"],
            "user_name": u["user_name"],
            "total": u["total"],
            "completed": u["completed"],
            "late": u["late"],
            "time_seconds": u["time_seconds"],
            "efficiency": eff,
        })
    user_performance.sort(key=lambda x: x["completed"], reverse=True)

    # Finanzas en el período (Estricto: Solo visible y accesible para Administradores)
    total_income = 0.0
    total_expenses = 0.0
    net_profit = 0.0

    if role_str == "administrador":
        fin_q = db.query(FinancialTransaction)
        if company_id:
            fin_q = fin_q.filter(FinancialTransaction.company_id == company_id)
        if project_id:
            fin_q = fin_q.filter(FinancialTransaction.project_id == project_id)
        if date_from:
            fin_q = fin_q.filter(FinancialTransaction.transaction_date >= date_from)
        if date_to:
            fin_q = fin_q.filter(FinancialTransaction.transaction_date <= date_to)

        fin_txs = fin_q.all()
        total_income = sum(float(t.amount or 0) for t in fin_txs if t.type == "ingreso")
        total_expenses = sum(float(t.amount or 0) for t in fin_txs if t.type == "egreso")
        net_profit = total_income - total_expenses


    # Lista formateada de actividades para la tabla del reporte (Top 100)
    activity_list = []
    for a in activities[:100]:
        st_val = a.status.value if hasattr(a.status, 'value') else str(a.status)
        tp_val = a.activity_type.value if hasattr(a.activity_type, 'value') else str(a.activity_type)
        pr_val = a.priority.value if hasattr(a.priority, 'value') else str(a.priority)
        activity_list.append({
            "id": a.id,
            "title": a.title,
            "activity_type": ACTIVITY_TYPE_LABELS.get(tp_val, tp_val),
            "priority": pr_val,
            "status": st_val,
            "status_label": ACTIVITY_STATUS_LABELS.get(st_val, st_val),
            "assigned_user": a.assigned_user.name if a.assigned_user else "Sin asignar",
            "project_name": a.project.name if a.project else "Sin Proyecto",
            "company_name": (a.project.company.name if a.project.company else "Sin Empresa") if a.project else "Sin Empresa",
            "deadline": str(a.deadline) if a.deadline else None,
            "time_spent_seconds": a.time_spent_seconds or 0,
            "created_at": str(a.created_at.date()) if a.created_at else None,
        })

    return {
        "kpis": {
            "total_activities": total_activities,
            "completed_activities": completed_activities,
            "pending_activities": pending_activities,
            "in_progress_activities": in_progress_activities,
            "in_review_activities": in_review_activities,
            "observed_activities": observed_activities,
            "cancelled_activities": cancelled_activities,
            "late_activities": late_activities,
            "total_time_seconds": total_time_seconds,
            "completion_rate": completion_rate,
            "total_income": round(total_income, 2) if role_str == "administrador" else None,
            "total_expenses": round(total_expenses, 2) if role_str == "administrador" else None,
            "net_profit": round(net_profit, 2) if role_str == "administrador" else None,
        },
        "activity_by_status": activity_by_status,
        "activity_by_type": activity_by_type,
        "user_performance": user_performance,
        "activities": activity_list,
    }



def export_activities_excel(db: Session, **filters) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = _get_activities_data(db, **filters)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actividades"

    headers = ["ID", "Título", "Tipo", "Prioridad", "Estado", "Fecha Límite", "Responsable", "Proyecto", "Empresa", "Fecha Creación"]
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    status_colors = {
        "aprobada": "22C55E", "en_proceso": "7C3AED", "en_revision": "3B82F6",
        "observada": "F59E0B", "cancelada": "EF4444", "pendiente": "9CA3AF", "asignada": "6366F1",
    }

    for r, row in enumerate(rows, 2):
        type_val = row.activity_type.value if hasattr(row.activity_type, "value") else str(row.activity_type)
        priority_val = row.priority.value if hasattr(row.priority, "value") else str(row.priority)
        status_val = row.status.value if hasattr(row.status, "value") else str(row.status)

        type_label = ACTIVITY_TYPE_LABELS.get(type_val, type_val)
        priority_label = PRIORITY_LABELS.get(priority_val, priority_val)
        status_label = ACTIVITY_STATUS_LABELS.get(status_val, status_val)

        data = [
            row.id,
            row.title,
            type_label,
            priority_label,
            status_label,
            str(row.deadline) if row.deadline else "",
            row.assigned_user or "-",
            row.project,
            row.company,
            str(row.created_at.date())
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            if col == 5:  # Columna de Estado
                color = status_colors.get(status_val, "9CA3AF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=actividades_{date.today()}.xlsx"},
    )


def export_activities_pdf(db: Session, **filters) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    rows = _get_activities_data(db, **filters)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=2*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], textColor=colors.HexColor("#7C3AED"), spaceAfter=12)

    story = [
        Paragraph("Reporte de Actividades", title_style),
        Paragraph(f"Generado: {date.today()}", styles["Normal"]),
        Spacer(1, 0.5*cm),
    ]

    headers = ["ID", "Título", "Tipo", "Estado", "Prioridad", "Responsable", "Proyecto", "Empresa", "Fecha Límite"]
    data = [headers]
    for row in rows:
        type_val = row.activity_type.value if hasattr(row.activity_type, "value") else str(row.activity_type)
        priority_val = row.priority.value if hasattr(row.priority, "value") else str(row.priority)
        status_val = row.status.value if hasattr(row.status, "value") else str(row.status)

        type_label = ACTIVITY_TYPE_LABELS.get(type_val, type_val)
        priority_label = PRIORITY_LABELS.get(priority_val, priority_val)
        status_label = ACTIVITY_STATUS_LABELS.get(status_val, status_val)

        data.append([
            str(row.id),
            row.title[:35],
            type_label,
            status_label,
            priority_label,
            row.assigned_user or "-",
            row.project[:20],
            row.company[:20],
            str(row.deadline) if row.deadline else "-",
        ])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C3AED")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=actividades_{date.today()}.pdf"},
    )
